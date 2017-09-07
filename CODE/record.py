from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import datetime
import numpy as np

FILENAME = '../System_Performance_Report.xlsx'

"""
# Example Here
Data_Collection = {}
Data_Collection["accuracy"] = 0.02
Data_Collection["threshold_score"] = "adadadaaaaaaa"
numberOfDataEntries = len(Data_Collection)

"""

def do_record(Data_Collection):
    try:
        wb = load_workbook(filename=FILENAME)
        SHEET_NAME = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        ws_new = wb.create_sheet(title=SHEET_NAME)

        # merge_cell
        ws_new.merge_cells('A1:F1')
        ws_new["A1"] = "Simulation Record"
        A1 = ws_new["A1"]
        A1.font = Font(bold=True, size=12)
        al1 = Alignment(horizontal='center', vertical='center')
        A1.alignment = al1

        ws_new.merge_cells('A2:A20')
        ws_new["A2"] = "Parameters Setting"
        A2 = ws_new["A2"]
        A2.font = Font(bold=True, size=12)
        al2 = Alignment(horizontal='center', vertical='center')
        A2.alignment = al2

        # load data to excel file in specific sheet page
        next_column = 2  # initial the start place of cell's column
        next_row = 2  # initial the start place of cell's column
        for key in Data_Collection:
            ws_new.cell(row=next_row, column=next_column, value=key)
            ws_new.cell(row=next_row, column=next_column + 1, value=Data_Collection[key])
            next_row += 1
        # save file
        wb.save(filename=FILENAME)
    except OSError:
        wb = Workbook()
        SHEET_NAME = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        ws_first = wb.active
        ws_first.title = SHEET_NAME

        # merge_cell
        ws_first.merge_cells('A1:F1')
        ws_first["A1"] = "Simulation Record"
        A1 = ws_first["A1"]
        A1.font = Font(bold=True, size=12)
        al1 = Alignment(horizontal='center', vertical='center')
        A1.alignment = al1

        ws_first.merge_cells('A2:A20')
        ws_first["A2"] = "Parameters Setting"
        A2 = ws_first["A2"]
        A2.font = Font(bold=True, size=12)
        al2 = Alignment(horizontal='center', vertical='center')
        A2.alignment = al2

        # load data to excel file in specific sheet page
        next_column = 2  # initial the start place of cell's column
        next_row = 2  # initial the start place of cell's column
        for key in Data_Collection:
            ws_first.cell(row=next_row, column=next_column, value=key)
            # print(Data_Collection[key])
            ws_first.cell(row=next_row, column=next_column + 1, value=Data_Collection[key])
            next_row += 1
        # save file
        wb.save(filename=FILENAME)